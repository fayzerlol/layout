"""
Streamlit application for managing and visualising equipment data.

This app loads an initial dataset extracted from the provided Excel workbook and
allows users to log in, assign qualitative scores, upload new data and view
dashboard metrics.  The interface has been customised with the company's
colour scheme and logo for a professional appearance.

To run the app locally, install the requirements listed in requirements.txt
and execute:

    streamlit run app.py

Files required in the same directory:
  - data.csv        : initial dataset with computed totals and statuses
  - download.png    : company logo to display in the header

  aa

"""

import os
import math
import pandas as pd
import numpy as np
import streamlit as st
import streamlit.components.v1 as components
import altair as alt
import openpyxl

DATA_FILE = 'data.csv'
LOGO_FILE = 'download.png'

# Simple authentication credentials.  In a real application these would be
# securely stored (e.g. environment variables or a credential manager).
USERS = {
    'admin': {'password': 'admin123', 'role': 'admin'},
    'user': {'password': 'user123', 'role': 'user'},
}

# -----------------------------------------------------------------------------
# Utility functions
# -----------------------------------------------------------------------------

@st.cache_data(show_spinner=False)
def load_data() -> pd.DataFrame:
    """Load the equipment dataset from disk into a DataFrame.

    The CSV must contain the following columns as produced by the data
    extraction script:
        - complexo, Local, TAG, Sistema, Modelo
        - PQS_Rec_Total, PQS_Inst_Total, Liq_Rec_Total, Liq_Inst_Total
        - Computed_Status
        - layout_sensores_cabo, layout_bicos_mangueiras, layout_painel_cabine
        - nota_prioridade_quantitativa, nota_prioridade_qualitativa
        - status_final_excel
        - nota_quantitativa, nota_qualitativa, status_final

    Returns:
        pd.DataFrame: Loaded dataset.
    """
    if not os.path.exists(DATA_FILE):
        raise FileNotFoundError(
            f"Data file '{DATA_FILE}' not found. Ensure it exists in the same directory.")
    df = pd.read_csv(DATA_FILE)
    return df


def save_data(df: pd.DataFrame) -> None:
    """Persist the DataFrame to disk.

    Args:
        df (pd.DataFrame): DataFrame to save.
    """
    df.to_csv(DATA_FILE, index=False)


def compute_totals(rec_pqs: str, inst_pqs: str, rec_liq: str, inst_liq: str):
    """Compute the numeric totals for PQS and liquid columns.

    The logic mirrors the original VBA `SomaCarga` function.  Strings of the
    form "2x60lbs + 1x30lbs" are parsed and converted into numeric values.

    Args:
        rec_pqs (str): Recommended PQS text.
        inst_pqs (str): Installed PQS text.
        rec_liq (str): Recommended liquid text.
        inst_liq (str): Installed liquid text.

    Returns:
        tuple[float, float, float, float]: Recommended PQS total, installed PQS total,
        recommended liquid total, installed liquid total.
    """

    def soma_carga(text: str, unidade: str) -> float:
        if not isinstance(text, str) or text.strip() == '':
            return 0.0
        text = text.replace(' ', '')
        total = 0.0
        for part in text.split('+'):
            if unidade in part and 'x' in part:
                try:
                    qty_str, value_str = part.split('x', 1)
                    qty = float(qty_str)
                    value = float(value_str.replace(unidade, ''))
                    total += qty * value
                except Exception:
                    continue
        return total

    rec_pqs_total = soma_carga(rec_pqs, 'lbs')
    inst_pqs_total = soma_carga(inst_pqs, 'lbs')
    rec_liq_total = soma_carga(rec_liq, 'gal')
    inst_liq_total = soma_carga(inst_liq, 'gal')
    return rec_pqs_total, inst_pqs_total, rec_liq_total, inst_liq_total


def compute_status(rec_pqs_tot: float, inst_pqs_tot: float,
                   rec_liq_tot: float, inst_liq_tot: float) -> str:
    """Compute the conformidade status based on recommended and installed totals.

    Args:
        rec_pqs_tot (float): Recommended PQS total.
        inst_pqs_tot (float): Installed PQS total.
        rec_liq_tot (float): Recommended liquid total.
        inst_liq_tot (float): Installed liquid total.

    Returns:
        str: 'CONFORME' or 'NÃO CONFORME'.
    """
    # Mirror the VBA logic of StatusCarga
    if rec_pqs_tot == 0 and rec_liq_tot == 0:
        return 'CONFORME'
    if (rec_pqs_tot > 0 and inst_pqs_tot == 0) or (rec_liq_tot > 0 and inst_liq_tot == 0):
        return 'NÃO CONFORME'
    if inst_pqs_tot >= rec_pqs_tot and inst_liq_tot >= rec_liq_tot:
        return 'CONFORME'
    return 'NÃO CONFORME'


def recalc_metrics(row: pd.Series) -> pd.Series:
    """Recalculate dependent fields based on qualitative and quantitative scores.

    Args:
        row (pd.Series): Row containing the scores.

    Returns:
        pd.Series: Row with updated fields (nota_quantitativa, nota_qualitativa, status_final).
    """
    # Compute quantitative note as the mean of the three layout scores where present
    scores = [row['layout_sensores_cabo'], row['layout_bicos_mangueiras'], row['layout_painel_cabine']]
    valid_scores = [s for s in scores if pd.notna(s)]
    if valid_scores:
        row['nota_quantitativa'] = round(sum(valid_scores) / len(valid_scores), 1)
    else:
        row['nota_quantitativa'] = np.nan
    # Qualitative note is taken directly from nota_prioridade_qualitativa
    row['nota_qualitativa'] = row.get('nota_qualitativa', row.get('nota_prioridade_qualitativa'))
    # Compute final status as sum of notes if both present
    if pd.notna(row['nota_quantitativa']) and pd.notna(row['nota_qualitativa']):
        row['status_final'] = round(row['nota_quantitativa'] + row['nota_qualitativa'], 1)
    elif pd.notna(row['nota_quantitativa']):
        row['status_final'] = row['nota_quantitativa']
    elif pd.notna(row['nota_qualitativa']):
        row['status_final'] = row['nota_qualitativa']
    else:
        row['status_final'] = np.nan
    return row


# -----------------------------------------------------------------------------
# Streamlit application
# -----------------------------------------------------------------------------

def main():
    st.set_page_config(page_title='Dashboard de Equipamentos', layout='wide')
    # Estilo customizado com gradiente, cards e visual profissional
    st.markdown(
        """
        <style>
        body, .reportview-container .main .block-container {
            background: linear-gradient(135deg, #232526 0%, #232526 100%);
            color: #f4f4f4;
        }
        .stButton>button {
            background: linear-gradient(90deg, #f4c247 0%, #00A85E 100%);
            color: #232526;
            font-weight: bold;
            border: none;
            padding: 0.6rem 1.2rem;
            border-radius: 8px;
            font-size: 1.1rem;
        }
        .stButton>button:hover {
            opacity: 0.85;
        }
        .status-cell {
            font-weight: bold;
            color: white;
            border-radius: 6px;
            padding: 4px 12px;
            font-size: 1rem;
            box-shadow: 0 2px 8px #23252633;
        }
        .status-conforme {
            background: linear-gradient(90deg, #00A85E 60%, #4fd953 100%);
        }
        .status-nc {
            background: linear-gradient(90deg, #d9534f 60%, #f44336 100%);
        }
        .status-aten {
            background: linear-gradient(90deg, #f4c247 60%, #ffe066 100%);
            color: #232526;
        }
        .card {
            background: #232526;
            border-radius: 12px;
            box-shadow: 0 2px 12px #00000022;
            padding: 1.2rem;
            margin-bottom: 1.2rem;
        }
        .table-prof {
            border-radius: 10px;
            overflow: hidden;
            font-size: 1rem;
        }
        th, td {
            padding: 8px 12px !important;
        }
        th {
            background: #2c2f33;
            color: #f4c247;
        }
        tr {
            background: #232526;
        }
        tr:nth-child(even) {
            background: #2c2f33;
        }
        </style>
        """,
        unsafe_allow_html=True
    )

    # Controle de login
    if 'logged_in' not in st.session_state:
        st.session_state.logged_in = False
        st.session_state.role = None
    if not st.session_state.logged_in:
        st.sidebar.image(LOGO_FILE, use_column_width=True)
        st.sidebar.title('Login')
        username = st.sidebar.text_input('Usuário')
        password = st.sidebar.text_input('Senha', type='password')
        if st.sidebar.button('Entrar'):
            user_info = USERS.get(username)
            if user_info and password == user_info['password']:
                st.session_state.logged_in = True
                st.session_state.role = user_info['role']
                st.experimental_rerun()
            else:
                st.sidebar.error('Usuário ou senha inválidos')
        st.stop()

    # Carregar dados
    df = load_data().copy()

    # Header com logo e título
    st.markdown('<div style="display:flex;align-items:center;gap:1.5rem;"><img src="download.png" style="height:60px;"> <h1 style="color:#f4c247;">Dashboard de Equipamentos</h1></div>', unsafe_allow_html=True)

    # Navegação lateral
    menu_options = ['Dashboard', 'Lançar Notas Qualitativas', 'Relatórios']
    if st.session_state.role == 'admin':
        menu_options.append('Admin: Importar Dados')
    choice = st.sidebar.radio('Menu', menu_options)

    # Dashboard principal
    if choice == 'Dashboard':
        with st.expander('Filtros', expanded=True):
            filtro_complexo = st.multiselect('Complexo', options=sorted(df['complexo'].dropna().unique()))
            filtro_local = st.multiselect('Local', options=sorted(df['Local'].dropna().unique()))
            filtro_sistema = st.multiselect('Sistema', options=sorted(df['Sistema'].dropna().unique()))
            filtro_fabricante = st.multiselect('Fabricante', options=sorted(df['Modelo'].dropna().unique()))
        filtered_df = df.copy()
        if filtro_complexo:
            filtered_df = filtered_df[filtered_df['complexo'].isin(filtro_complexo)]
        if filtro_local:
            filtered_df = filtered_df[filtered_df['Local'].isin(filtro_local)]
        if filtro_sistema:
            filtered_df = filtered_df[filtered_df['Sistema'].isin(filtro_sistema)]
        if filtro_fabricante:
            filtered_df = filtered_df[filtered_df['Modelo'].isin(filtro_fabricante)]

        # Cards de métricas principais
        m1, m2, m3 = st.columns(3)
        status_counts = filtered_df['Computed_Status'].value_counts().to_dict()
        conforme = status_counts.get('CONFORME', 0)
        nao_conforme = status_counts.get('NÃO CONFORME', 0)
        pendentes = filtered_df[filtered_df['nota_qualitativa'].isna() | filtered_df['nota_quantitativa'].isna()]
        m1.metric('Equipamentos CONFORME', conforme)
        m2.metric('Equipamentos NÃO CONFORME', nao_conforme)
        m3.metric('Pendentes de Nota', len(pendentes))

        # 1. Descrição dos equipamentos
        st.markdown('<div class="card"><h2 style="color:#f4c247;">1. Descrição dos Equipamentos</h2>' +
            filtered_df[['complexo', 'Local', 'TAG', 'Sistema', 'Modelo']].to_html(classes='table-prof', index=False) + '</div>', unsafe_allow_html=True)

        # 2. Status dos agentes extintores
        def status_cell(val):
            if val == 'CONFORME':
                return '<span class="status-cell status-conforme">✔ CONFORME</span>'
            elif val == 'NÃO CONFORME':
                return '<span class="status-cell status-nc">✖ NÃO CONFORME</span>'
            else:
                return val
        extintores_df = filtered_df[['TAG', 'PQS_Rec_Total', 'PQS_Inst_Total', 'Liq_Rec_Total', 'Liq_Inst_Total', 'Computed_Status']].copy()
        extintores_df['Status'] = extintores_df['Computed_Status'].apply(status_cell)
        st.markdown('<div class="card"><h2 style="color:#f4c247;">2. Status dos Agentes Extintores</h2>' +
            extintores_df.to_html(escape=False, classes='table-prof', index=False) + '</div>', unsafe_allow_html=True)

        # 2.1 Modal para PDF do relatório
        st.markdown('<div class="card"><h2 style="color:#f4c247;">Relatório PDF do Equipamento</h2>', unsafe_allow_html=True)
        tags_with_reports = filtered_df.dropna(subset=['report_id'])[['TAG', 'report_id']]
        if not tags_with_reports.empty:
            selected_row = st.selectbox('Selecione o equipamento (TAG):', tags_with_reports['TAG'])
            report_id = tags_with_reports[tags_with_reports['TAG'] == selected_row]['report_id'].values[0]
            file_id = str(report_id).strip()
            if '_' in file_id:
                file_id = file_id.split('_')[-1]
            if '.' in file_id:
                file_id = file_id.split('.')[0]
            url = f'https://drive.google.com/file/d/{file_id}/preview'
            st.markdown(f'<iframe src="{url}" width="900" height="500" style="border-radius:12px;border:2px solid #f4c247;"></iframe>', unsafe_allow_html=True)
        else:
            st.info('Nenhum relatório associado aos equipamentos selecionados.')
        st.markdown('</div>', unsafe_allow_html=True)

        # 3. Sensores
        st.markdown('<div class="card"><h2 style="color:#f4c247;">3. Sensores: Recomendado vs Instalado</h2>' +
            filtered_df[['TAG', 'layout_sensores_cabo']].to_html(classes='table-prof', index=False) + '</div>', unsafe_allow_html=True)

        # 4. Bicos PQS
        st.markdown('<div class="card"><h2 style="color:#f4c247;">4. Bicos PQS: Recomendado vs Instalado</h2>' +
            filtered_df[['TAG', 'layout_bicos_mangueiras']].to_html(classes='table-prof', index=False) + '</div>', unsafe_allow_html=True)

        # 5. Bicos Líquidos
        st.markdown('<div class="card"><h2 style="color:#f4c247;">5. Bicos Líquidos: Recomendado vs Instalado</h2>' +
            filtered_df[['TAG', 'layout_painel_cabine']].to_html(classes='table-prof', index=False) + '</div>', unsafe_allow_html=True)

        # 6. Notas Qualitativas e Quantitativas
        st.markdown('<div class="card"><h2 style="color:#f4c247;">6. Notas Qualitativas e Quantitativas</h2>' +
            filtered_df[['TAG', 'nota_quantitativa', 'nota_qualitativa', 'status_final']].to_html(classes='table-prof', index=False) + '</div>', unsafe_allow_html=True)

        # 7. Equipamentos Pendentes de Nota
        st.markdown('<div class="card"><h2 style="color:#f4c247;">7. Equipamentos Pendentes de Nota</h2>' +
            pendentes[['TAG', 'nota_quantitativa', 'nota_qualitativa']].to_html(classes='table-prof', index=False) + '</div>', unsafe_allow_html=True)

        # 8. Observações dos Equipamentos
        obs_col = 'observacao' if 'observacao' in filtered_df.columns else None
        if obs_col:
            st.markdown('<div class="card"><h2 style="color:#f4c247;">8. Observações dos Equipamentos</h2>' +
                filtered_df[['TAG', obs_col]].to_html(classes='table-prof', index=False) + '</div>', unsafe_allow_html=True)
        else:
            st.info('Coluna de observação não encontrada.')

        # 9. Notas dos Campos Comparativos
        st.markdown('<div class="card"><h2 style="color:#f4c247;">9. Notas dos Campos Comparativos</h2>' +
            filtered_df[['TAG', 'nota_quantitativa', 'nota_qualitativa']].to_html(classes='table-prof', index=False) + '</div>', unsafe_allow_html=True)

        # 10. Status Final com Gradiente de Cor
        def gradiente_status(val):
            if pd.isna(val):
                return '<span class="status-cell">N/A</span>'
            if val <= 4:
                return '<span class="status-cell status-nc">{}</span>'.format(val)
            elif val <= 7:
                return '<span class="status-cell status-aten">{}</span>'.format(val)
            else:
                return '<span class="status-cell status-conforme">{}</span>'.format(val)
        grad_df = filtered_df[['TAG', 'status_final']].copy()
        grad_df['Status Final'] = grad_df['status_final'].apply(gradiente_status)
        st.markdown('<div class="card"><h2 style="color:#f4c247;">10. Status Final com Gradiente de Cor</h2>' +
            grad_df[['TAG', 'Status Final']].to_html(escape=False, classes='table-prof', index=False) + '</div>', unsafe_allow_html=True)

    # Página de lançamento de notas qualitativas
    elif choice == 'Lançar Notas Qualitativas':
        st.subheader('Inserir/Atualizar Notas Qualitativas')
        selected_tag = st.selectbox('Selecione um equipamento (TAG):', df['TAG'])
        row_idx = df[df['TAG'] == selected_tag].index[0]
        row = df.loc[row_idx]
        st.markdown(f"**Complexo:** {row['complexo']} | **Local:** {row['Local']} | **Status carga:** {row['Computed_Status']}")
        def_input = lambda x: 0.0 if pd.isna(x) else float(x)
        val1 = st.slider('Layout Monitoramento Sensores/Cabo Linear (1-5)', 0.0, 5.0, def_input(row['layout_sensores_cabo']), 0.5)
        val2 = st.slider('Layout Bicos e Mangueiras (1-5)', 0.0, 5.0, def_input(row['layout_bicos_mangueiras']), 0.5)
        val3 = st.slider('Layout Posicionamento Acionadores e Painel da Cabine (1-5)', 0.0, 5.0, def_input(row['layout_painel_cabine']), 0.5)
        qual = st.slider('Nota Qualitativa (1-5)', 0.0, 5.0, def_input(row.get('nota_qualitativa')), 0.5)
        observacao = st.text_area('Observação (obrigatório)', value=row.get('observacao', ''))
        if st.button('Salvar Notas'):
            if not (observacao or '').strip():
                st.error('Campo de observação é obrigatório!')
            else:
                df.at[row_idx, 'layout_sensores_cabo'] = val1 if val1 > 0 else np.nan
                df.at[row_idx, 'layout_bicos_mangueiras'] = val2 if val2 > 0 else np.nan
                df.at[row_idx, 'layout_painel_cabine'] = val3 if val3 > 0 else np.nan
                df.at[row_idx, 'nota_qualitativa'] = qual if qual > 0 else np.nan
                df.at[row_idx, 'observacao'] = observacao
                df.loc[row_idx] = recalc_metrics(df.loc[row_idx])
                save_data(df)
                st.success('Notas salvas com sucesso!')

    # Relatórios page
    elif choice == 'Relatórios':
        st.subheader('Relatórios de Equipamentos')
        # Filters for reports
        with st.expander('Filtros'):
            filtro_complexo = st.multiselect('Complexo', options=sorted(df['complexo'].dropna().unique()))
            filtro_local = st.multiselect('Local', options=sorted(df['Local'].dropna().unique()))
            filtro_sistema = st.multiselect('Sistema', options=sorted(df['Sistema'].dropna().unique()))
            filtro_fabricante = st.multiselect('Modelo', options=sorted(df['Modelo'].dropna().unique()))
        filtered_df = df.copy()
        if filtro_complexo:
            filtered_df = filtered_df[filtered_df['complexo'].isin(filtro_complexo)]
        if filtro_local:
            filtered_df = filtered_df[filtered_df['Local'].isin(filtro_local)]
        if filtro_sistema:
            filtered_df = filtered_df[filtered_df['Sistema'].isin(filtro_sistema)]
        if filtro_fabricante:
            filtered_df = filtered_df[filtered_df['Modelo'].isin(filtro_fabricante)]
        # Display table
        display_cols = ['complexo', 'Local', 'TAG', 'Sistema', 'Modelo', 'Computed_Status',
                        'nota_quantitativa', 'nota_qualitativa', 'status_final', 'report_id']
        st.dataframe(filtered_df[display_cols], use_container_width=True)
        # Select report to view
        st.markdown('### Visualizar relatório em PDF')
        tags_with_reports = filtered_df.dropna(subset=['report_id'])[['TAG', 'report_id']]
        if not tags_with_reports.empty:
            selected_row = st.selectbox('Selecione o equipamento (TAG):', tags_with_reports['TAG'])
            report_id = tags_with_reports[tags_with_reports['TAG'] == selected_row]['report_id'].values[0]
            # If report_id contains file extension, split by '.' or '_' and take last part
            file_id = str(report_id).strip()
            # Many file names end with the id separated by '_'; attempt to extract trailing segment
            if '_' in file_id:
                file_id = file_id.split('_')[-1]
            if '.' in file_id:
                file_id = file_id.split('.')[0]
            # Build Google Drive preview URL
            url = f'https://drive.google.com/file/d/{file_id}/preview'
            components.iframe(url, width=700, height=500)
        else:
            st.info('Nenhum relatório associado aos equipamentos selecionados.')
    # Admin upload page
    elif choice == 'Admin: Importar Dados' and st.session_state.role == 'admin':
        st.subheader('Importar/Acrescentar Dados')
        st.write('Carregue um arquivo Excel (.xlsx/.xlsm) ou CSV contendo dados de equipamentos.\n'
                 'As colunas relevantes serão extraídas automaticamente. Registros existentes com o mesmo TAG serão atualizados.')
        uploaded_file = st.file_uploader('Selecione o arquivo', type=['xlsx', 'xlsm', 'csv'])
        if uploaded_file is not None:
            try:
                # Determine type and read accordingly
                if uploaded_file.name.lower().endswith(('.xlsx', '.xlsm')):
                    # Read the sheet named 'VALE - MG' or first sheet
                    wb = openpyxl.load_workbook(uploaded_file, data_only=True)
                    sheet_name = 'VALE - MG' if 'VALE - MG' in wb.sheetnames else wb.sheetnames[0]
                    ws_new = wb[sheet_name]
                    # Build DataFrame similarly to the extraction script
                    header_row_new = 6
                    headers_new = [ws_new.cell(header_row_new, col).value for col in range(1, ws_new.max_column+1)]
                    # Map relevant columns
                    indices = {}
                    desired = ['complexo', 'Local', 'TAG', 'Sistema', 'Modelo',
                               'Recomendado\nQtd. PQS', 'Instalado\nQtd. PQS',
                               'Recomendado\nQTD. Ag Líquido', 'Instalado\nQTD. Ag Líquido',
                               '5.)Layout Monitoramento Sensores/Cabo Linear',
                               '6.) Layout Bicos e Mangueiras',
                               '7.)Layout Posicionamento Acionadores e Painel da Cabine',
                               'Nota de Prioridade Qualitativa(NOTAS: 1 PIOR NOTA;  5 MELHOR NOTA)']
                    for col_idx, header in enumerate(headers_new, start=1):
                        if header in desired:
                            indices[header] = col_idx
                    # Extract rows
                    new_rows = []
                    for r in range(header_row_new+1, ws_new.max_row+1):
                        row = {}
                        for header, idx in indices.items():
                            row[header] = ws_new.cell(r, idx).value
                        new_rows.append(row)
                    new_df = pd.DataFrame(new_rows)
                else:
                    # CSV
                    new_df = pd.read_csv(uploaded_file)
                # Ensure TAG exists
                if 'TAG' not in new_df.columns or new_df['TAG'].isna().all():
                    st.error('O arquivo importado não contém coluna "TAG" válida.')
                else:
                    # Compute totals and status for new rows
                    def parse_score(s):
                        try:
                            return float(str(s).replace(',', '.'))
                        except:
                            return np.nan
                    # Process each row
                    processed = []
                    for _, record in new_df.iterrows():
                        rec_pqs = record.get('Recomendado\nQtd. PQS')
                        inst_pqs = record.get('Instalado\nQtd. PQS')
                        rec_liq = record.get('Recomendado\nQTD. Ag Líquido')
                        inst_liq = record.get('Instalado\nQTD. Ag Líquido')
                        rec_pqs_tot, inst_pqs_tot, rec_liq_tot, inst_liq_tot = compute_totals(str(rec_pqs), str(inst_pqs), str(rec_liq), str(inst_liq))
                        status = compute_status(rec_pqs_tot, inst_pqs_tot, rec_liq_tot, inst_liq_tot)
                        # Build row dict
                        row_dict = {
                            'complexo': record.get('complexo'),
                            'Local': record.get('Local'),
                            'TAG': record.get('TAG'),
                            'Sistema': record.get('Sistema'),
                            'Modelo': record.get('Modelo'),
                            'PQS_Rec_Total': rec_pqs_tot,
                            'PQS_Inst_Total': inst_pqs_tot,
                            'Liq_Rec_Total': rec_liq_tot,
                            'Liq_Inst_Total': inst_liq_tot,
                            'Computed_Status': status,
                            'layout_sensores_cabo': parse_score(record.get('5.)Layout Monitoramento Sensores/Cabo Linear')),
                            'layout_bicos_mangueiras': parse_score(record.get('6.) Layout Bicos e Mangueiras')),
                            'layout_painel_cabine': parse_score(record.get('7.)Layout Posicionamento Acionadores e Painel da Cabine')),
                            'nota_qualitativa': parse_score(record.get('Nota de Prioridade Qualitativa(NOTAS: 1 PIOR NOTA;  5 MELHOR NOTA)')),
                        }
                        # Calculate quantitative and final notes
                        temp_series = pd.Series(row_dict)
                        temp_series = recalc_metrics(temp_series)
                        processed.append(temp_series.to_dict())
                    processed_df = pd.DataFrame(processed)
                    # Merge with existing df on TAG (update existing rows)
                    df_update = df.set_index('TAG')
                    processed_update = processed_df.set_index('TAG')
                    df_update.update(processed_update)
                    # Append new rows that do not exist
                    new_tags = processed_update.index.difference(df_update.index)
                    if len(new_tags) > 0:
                        df_append = processed_update.loc[new_tags]
                        df_update = pd.concat([df_update, df_append])
                    # Reset index
                    df_update = df_update.reset_index()
                    save_data(df_update)
                    st.success('Dados importados com sucesso!')
            except Exception as e:
                st.error(f'Erro ao importar arquivo: {e}')

    # Footer
    st.sidebar.markdown('---')
    st.sidebar.markdown('Desenvolvido para Grupo Franzen')


if __name__ == '__main__':
    main()